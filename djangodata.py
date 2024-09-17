import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import re

excel_file_path = 'C:/Users/User/Desktop/Inventory.xlsx'
wb = load_workbook(excel_file_path)
sheet = wb.active

url = 'https://example.com/newproduct/'

session = requests.Session()

response = session.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

csrf_token = soup.find('input', {'name': 'csrfmiddlewaretoken'})['value']

categories = {option.text.lower(): option['value'] for option in soup.select('select[name="product_category"] option') if option['value']}

def clean_numeric_value(value, expected_type=float):
    if isinstance(value, (int, float)):
        return value
    try:
        numeric_value = re.findall(r"[-+]?\d*\.\d+|\d+", str(value))
        if numeric_value:
            return expected_type(numeric_value[0])
        else:
            raise ValueError(f"Could not find numeric value in {value}")
    except ValueError as e:
        raise ValueError(f"Invalid numeric value: {value}. Error: {e}")

for row in sheet.iter_rows(min_row=2, values_only=True):  
    product_name = row[0]  
    category = row[1]     
    cost = row[2]         
    price = row[3]        
    quantity = row[4]     
    serial_number = row[5] 
    description = row[6]   

    try:
        cost = clean_numeric_value(cost, float)
        price = clean_numeric_value(price, float)

        quantity = clean_numeric_value(quantity, int)
    except ValueError as e:
        print(f"Skipping invalid row due to error: {e}")
        continue

    if category.lower() in categories:
        product_category = categories[category.lower()]
        new_category_name = None
    else:
        product_category = 'new'
        new_category_name = category

    data = {
        'csrfmiddlewaretoken': csrf_token,
        'product_name': product_name,
        'product_category': product_category,
        'cost': cost,
        'price': price,
        'quantity': quantity,
        'serial_number': serial_number,
        'description': description
    }

    if new_category_name:
        data['new_category_name'] = new_category_name

    headers = {
        'Referer': url,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36'
    }

    response = session.post(url, data=data, headers=headers)

    if response.status_code == 200:
        print(f'Successfully added product: {product_name}')
    elif response.status_code == 500:
        print(f'Server error while adding product: {product_name}')
    else:
        print(f'Failed to add product: {product_name}, Status code: {response.status_code}')
